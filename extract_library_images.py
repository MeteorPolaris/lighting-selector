from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader


DEFAULT_EXCEL = Path("产品自动选型工具_1 Leon&Yawen_室外路灯泛光灯-2 图片库.xlsx")
DEFAULT_SHEET = "PHOTO Database"
DEFAULT_OUTPUT_DIR = Path("assets")
DEFAULT_IMAGE_COLS = ("C", "D")


def sanitize_filename(name: str) -> str:
    cleaned = name.strip()
    cleaned = re.sub(r"[\\\\/:*?\"<>|]", "_", cleaned)
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "unknown"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从图片库 Excel 中按 A 列型号名导出行内图片到 assets/ 目录。"
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL,
        help=f"图片库 Excel 路径，默认: {DEFAULT_EXCEL}",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=DEFAULT_SHEET,
        help=f"工作表名称，默认: {DEFAULT_SHEET}",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"输出目录，默认: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="起始行（含），默认 2",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=0,
        help="结束行（含），默认 0 表示到最后一行",
    )
    parser.add_argument(
        "--cols",
        type=str,
        default=",".join(DEFAULT_IMAGE_COLS),
        help="图片列，逗号分隔，默认 C,D",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_path = args.excel.resolve()
    output_dir = args.out.resolve()
    image_cols = tuple(col.strip().upper() for col in args.cols.split(",") if col.strip())

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = load_workbook(excel_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {args.sheet}，可用工作表: {wb.sheetnames}")

    ws = wb[args.sheet]
    image_loader = SheetImageLoader(ws)
    output_dir.mkdir(parents=True, exist_ok=True)

    start_row = max(1, args.start_row)
    end_row = ws.max_row if args.end_row <= 0 else min(args.end_row, ws.max_row)

    saved_count = 0
    skipped_no_name = 0
    skipped_no_image = 0

    for row in range(start_row, end_row + 1):
        model_value = ws[f"A{row}"].value
        if model_value is None:
            skipped_no_name += 1
            continue

        model_name = sanitize_filename(str(model_value))
        row_has_image = False

        for col in image_cols:
            cell = f"{col}{row}"
            if not image_loader.image_in(cell):
                continue

            image = image_loader.get(cell)
            suffix = f"_{col.lower()}" if len(image_cols) > 1 else ""
            output_path = output_dir / f"{model_name}{suffix}.png"
            image.save(output_path)
            row_has_image = True
            saved_count += 1
            print(f"Saved: {output_path}")

        if not row_has_image:
            skipped_no_image += 1

    print("\nDone.")
    print(f"Sheet: {args.sheet}")
    print(f"Rows scanned: {end_row - start_row + 1}")
    print(f"Images saved: {saved_count}")
    print(f"Rows skipped (A列空): {skipped_no_name}")
    print(f"Rows skipped (无图片): {skipped_no_image}")


if __name__ == "__main__":
    main()
